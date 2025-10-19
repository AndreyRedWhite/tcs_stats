# json_to_exel.py
from __future__ import annotations
import json
from pathlib import Path
from typing import Dict, Any, List

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def _autosize(ws: Worksheet):
    for col_cells in ws.columns:
        max_len = 0
        col = col_cells[0].column if hasattr(col_cells[0], "column") else col_cells[0].column_letter
        for c in col_cells:
            try:
                v = "" if c.value is None else str(c.value)
                max_len = max(max_len, len(v))
            except Exception:
                pass
        adj = max_len + 2
        ws.column_dimensions[get_column_letter(col)].width = min(adj, 60)


def _overview_from_windows(windows: Dict[str, Any]) -> pd.DataFrame:
    rows = []
    for win_name, payload in windows.items():
        stats: Dict[str, Any] = payload.get("stats", {})
        # агрегируем по окну
        tot_trades = 0
        tot_roundtrips = 0
        wins = 0
        losses = 0
        pnl_gross = 0.0
        commissions = 0.0
        pnl_net = 0.0
        for _, s in stats.items():
            tot_trades += s.get("n_trades", 0)
            tot_roundtrips += s.get("n_roundtrips", 0)
            wins += s.get("wins", 0)
            losses += s.get("losses", 0)
            pnl_gross += float(s.get("pnl_gross", 0.0))
            commissions += float(s.get("commissions", 0.0))
            pnl_net += float(s.get("pnl_net", 0.0))
        winrate = (wins / (wins + losses) * 100.0) if (wins + losses) > 0 else 0.0
        rows.append({
            "window": win_name,
            "from": payload.get("from"),
            "to": payload.get("to"),
            "trades": tot_trades,
            "roundtrips": tot_roundtrips,
            "wins": wins,
            "losses": losses,
            "winrate_%": round(winrate, 2),
            "pnl_gross": round(pnl_gross, 2),
            "commissions": round(commissions, 2),
            "pnl_net": round(pnl_net, 2),
        })
    return pd.DataFrame(rows).sort_values("window")


def _instruments_table(win_name: str, payload: Dict[str, Any]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for _, s in payload.get("stats", {}).items():
        wins = s.get("wins", 0)
        losses = s.get("losses", 0)
        wr = (wins / (wins + losses) * 100.0) if (wins + losses) > 0 else 0.0
        rows.append({
            "ticker": s.get("ticker"),
            "class_code": s.get("class_code"),
            "figi": s.get("figi"),
            "uid": s.get("uid"),
            "n_trades": s.get("n_trades"),
            "n_roundtrips": s.get("n_roundtrips"),
            "wins": wins,
            "losses": losses,
            "winrate_%": round(wr, 2),
            "pnl_gross": s.get("pnl_gross"),
            "commissions": s.get("commissions"),
            "pnl_net": s.get("pnl_net"),
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["pnl_net", "pnl_gross"], ascending=False)
    return df


def _roundtrips_table(win_name: str, payload: Dict[str, Any]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for _, s in payload.get("stats", {}).items():
        ticker = s.get("ticker")
        for rt in s.get("roundtrips", []):
            rows.append({
                "ticker": ticker,
                "open_time": rt.get("open_time"),
                "close_time": rt.get("close_time"),
                "qty": rt.get("qty"),
                "buy_price": rt.get("buy_price"),
                "sell_price": rt.get("sell_price"),
                "pnl_gross": rt.get("pnl_gross"),
                "is_win": rt.get("is_win"),
            })
    df = pd.DataFrame(rows)
    if not df.empty:
        df = df.sort_values(["ticker", "close_time"])
    return df


def export_to_excel(src_json: Path, out_xlsx: Path):
    with open(src_json, "r", encoding="utf-8") as f:
        data = json.load(f)

    windows: Dict[str, Any] = data.get("windows", {})
    if not windows:
        raise SystemExit("В JSON нет секции 'windows' — нечего экспортировать.")

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        # Overview
        df_over = _overview_from_windows(windows)
        df_over.to_excel(writer, sheet_name="Overview", index=False)

        # По окнам — свод по инструментам и «сделки-связки» (roundtrips)
        for win_name, payload in windows.items():
            df_instr = _instruments_table(win_name, payload)
            df_rt = _roundtrips_table(win_name, payload)

            ws_name_instr = f"Instruments_{win_name}"
            ws_name_rt = f"RoundTrips_{win_name}"

            if not df_instr.empty:
                df_instr.to_excel(writer, sheet_name=ws_name_instr, index=False)
            else:
                pd.DataFrame([{"note": "Нет сделок в этом окне"}]).to_excel(
                    writer, sheet_name=ws_name_instr, index=False
                )

            if not df_rt.empty:
                df_rt.to_excel(writer, sheet_name=ws_name_rt, index=False)
            else:
                pd.DataFrame([{"note": "Нет закрытых раундов в этом окне"}]).to_excel(
                    writer, sheet_name=ws_name_rt, index=False
                )

        # красота: автоширина, автофильтр, freeze panes
        book = writer.book
        for ws in book.worksheets:
            # автофильтр на первую строку
            if ws.max_row >= 1 and ws.max_column >= 1:
                ws.auto_filter.ref = ws.dimensions
            # заморозить заголовок
            ws.freeze_panes = "A2"
            # ширина
            _autosize(ws)

    print(f"Готово: {out_xlsx}")


if __name__ == "__main__":
    import argparse
    p = argparse.ArgumentParser(description="Экспорт tcs_stats.json → Excel")
    p.add_argument("src", nargs="?", default="tcs_stats.json", help="Путь к JSON (по умолчанию tcs_stats.json)")
    p.add_argument("-o", "--out", default="tcs_stats.xlsx", help="Путь к Excel файлу")
    args = p.parse_args()

    export_to_excel(Path(args.src), Path(args.out))
